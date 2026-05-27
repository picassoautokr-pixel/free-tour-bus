# -*- coding: utf-8 -*-
"""Export docs/quote_support_field_map.csv -> docs/quote_support_field_map.xlsx."""
from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "docs" / "quote_support_field_map.csv"
XLSX_PATH = ROOT / "docs" / "quote_support_field_map.xlsx"
SHEET_NAME = "quote_support_field_map"

# 검증용 한글 샘플 (헤더·대표 행에 포함되어야 함)
KOREAN_MARKERS = ("구분", "테이블명", "희망 일반견적", "지원금", "견적")


def has_korean(text: str) -> bool:
    return bool(re.search(r"[\uac00-\ud7a3]", text))


def load_dataframe() -> pd.DataFrame:
    if not CSV_PATH.is_file():
        raise FileNotFoundError(f"CSV not found: {CSV_PATH}")
    df = pd.read_csv(CSV_PATH, encoding="utf-8", dtype=str).fillna("")
    return df


def auto_column_widths(ws, max_width: int = 60) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            lines = str(value).splitlines()
            line_len = max((len(line) for line in lines), default=0)
            max_len = max(max_len, line_len)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def apply_sheet_format(ws) -> None:
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = wrap

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    auto_column_widths(ws)


def verify_korean(df: pd.DataFrame) -> None:
    header = "".join(df.columns.astype(str))
    if not has_korean(header):
        raise ValueError("Header row missing Korean characters")

    sample = df.head(10).astype(str).values.flatten()
    joined = "".join(sample)
    if not any(m in joined for m in KOREAN_MARKERS):
        raise ValueError("Sample data missing expected Korean markers")

    for col in ("한글명", "역할정의", "표시 라벨"):
        if col not in df.columns:
            raise ValueError(f"Missing column: {col}")
        if not df[col].astype(str).str.contains(r"[\uac00-\ud7a3]", regex=True).any():
            raise ValueError(f"Column {col} has no Korean text")


def export_xlsx(df: pd.DataFrame) -> None:
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(
        XLSX_PATH,
        engine="openpyxl",
    ) as writer:
        df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        ws = writer.sheets[SHEET_NAME]
        apply_sheet_format(ws)


def print_first_rows(df: pd.DataFrame, n: int = 5) -> None:
    print(f"--- Korean verification: first {n} rows ---")
    cols = list(df.columns)
    for i in range(min(n, len(df))):
        row = df.iloc[i]
        print(f"[row {i + 1}]")
        for c in cols[:6]:
            print(f"  {c}: {row[c]}")
        print(f"  한글명: {row.get('한글명', '')}")
        print()


def write_verification_log(df: pd.DataFrame, n: int = 5) -> Path:
    log_path = XLSX_PATH.parent / "_xlsx_korean_verify.txt"
    lines = [
        f"file: {XLSX_PATH}",
        f"sheet: {SHEET_NAME}",
        f"rows: {len(df)}, columns: {len(df.columns)}",
        f"headers: {list(df.columns)}",
        "",
    ]
    for i in range(min(n, len(df))):
        row = df.iloc[i]
        lines.append(f"[row {i + 1}]")
        lines.append(f"  구분: {row['구분']}")
        lines.append(f"  테이블명: {row['테이블명']}")
        lines.append(f"  필드명: {row['필드명']}")
        lines.append(f"  한글명: {row['한글명']}")
        lines.append(f"  역할정의: {str(row['역할정의'])[:80]}")
        lines.append("")
    log_path.write_text("\n".join(lines), encoding="utf-8")
    return log_path


def main() -> int:
    df = load_dataframe()
    verify_korean(df)
    export_xlsx(df)

    # 저장 후 재읽기 검증
    df2 = pd.read_excel(XLSX_PATH, sheet_name=SHEET_NAME, engine="openpyxl", dtype=str).fillna(
        ""
    )
    verify_korean(df2)

    log_path = write_verification_log(df2, 5)
    print(f"Saved: {XLSX_PATH}")
    print(f"Rows: {len(df2)}, Columns: {len(df2.columns)}")
    print(f"Korean verify log: {log_path}")
    print_first_rows(df2, 5)
    return 0


if __name__ == "__main__":
    sys.exit(main())
