# -*- coding: utf-8 -*-
"""Generate docs/*.xlsx workbooks (field_map, state_definition, calculation_rules, dashboard_display_rules)."""
from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"

FIELD_MAP_CSV = DOCS / "quote_support_field_map.csv"
FIELD_MAP_XLSX = DOCS / "field_map.xlsx"
STATE_XLSX = DOCS / "state_definition.xlsx"
CALC_XLSX = DOCS / "calculation_rules.xlsx"
DISPLAY_XLSX = DOCS / "dashboard_display_rules.xlsx"

KOREAN_RE = re.compile(r"[\uac00-\ud7a3]")


def has_korean(text: str) -> bool:
    return bool(KOREAN_RE.search(text))


def auto_column_widths(ws, max_width: int = 72) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            lines = str(cell.value).splitlines()
            max_len = max(max_len, max((len(line) for line in lines), default=0))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def apply_sheet_format(ws) -> None:
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = wrap
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    auto_column_widths(ws)


def save_workbook(path: Path, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        apply_sheet_format(writer.sheets[sheet_name])
    return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", dtype=str).fillna("")


def verify_dataframe_korean(df: pd.DataFrame, label_cols: list[str]) -> None:
    header = "".join(df.columns.astype(str))
    if not has_korean(header):
        raise ValueError(f"Header missing Korean: {list(df.columns)}")
    for col in label_cols:
        if col not in df.columns:
            raise ValueError(f"Missing column: {col}")
    sample = df.head(5).astype(str)
    if not sample.apply(lambda s: s.str.contains(KOREAN_RE, regex=True)).any().any():
        raise ValueError("First 5 rows contain no Korean in any column")


def print_verify(path: Path, sheet: str, df: pd.DataFrame, preview_cols: list[str], n: int = 5) -> None:
    print(f"\n=== {path.name} ({len(df)} rows) ===")
    verify_dataframe_korean(df, preview_cols)
    for i in range(min(n, len(df))):
        row = df.iloc[i]
        parts = [f"{c}={row[c]}" for c in preview_cols if c in df.columns]
        print(f"  [{i + 1}] " + " | ".join(parts[:4]))


def build_field_map() -> pd.DataFrame:
    if not FIELD_MAP_CSV.is_file():
        raise FileNotFoundError(FIELD_MAP_CSV)
    df = pd.read_csv(FIELD_MAP_CSV, encoding="utf-8", dtype=str).fillna("")
    if "데이터 타입" in df.columns:
        df = df.rename(columns={"데이터 타입": "데이터타입"})
    expected = [
        "구분",
        "테이블명",
        "필드명",
        "한글명",
        "역할정의",
        "데이터타입",
        "값 예시",
        "계산식",
        "저장 시점",
        "수정 주체",
        "사용하는 API",
        "사용하는 파일/컴포넌트",
        "표시되는 대시보드",
        "표시 라벨",
        "fallback 우선순위",
        "주의사항",
    ]
    for col in expected:
        if col not in df.columns:
            df[col] = ""
    return df[expected]


def build_state_definition() -> pd.DataFrame:
    rows = [
        # 클라이언트 탭 (필수)
        (
            "클라이언트 탭",
            "requesting",
            "견적요청중",
            "견적 수집·마감 전 신청",
            "final_selected_quote_id 없음 AND quote_status가 자동마감 상태가 아님",
            "자동마감 상태 전이 또는 매칭 확정",
            "시스템·고객",
            "applications",
            "quote_status; final_selected_quote_id",
            "클라이언트",
            "자동마감 또는 매칭완료",
            "clientApplicationTab()",
        ),
        (
            "클라이언트 탭",
            "auto_closed",
            "자동마감",
            "마감됐으나 고객 매칭 전",
            "quote_status ∈ {closed_by_time, closed_by_quote_count, closed_by_price, manually_closed, auto_selected, extended_no_quotes} AND final_selected_quote_id 없음",
            "고객 매칭 확정",
            "시스템·고객",
            "applications",
            "quote_status; final_selected_quote_id",
            "클라이언트",
            "매칭완료",
            "AUTO_CLOSED_STATUSES",
        ),
        (
            "클라이언트 탭",
            "matched",
            "매칭완료",
            "고객이 견적 선택·매칭 확정",
            "final_selected_quote_id 존재",
            "출발일 경과 시 진행완료 서브필터",
            "고객",
            "applications",
            "final_selected_quote_id; final_selected_quote_source; selected_price_*",
            "클라이언트",
            "(종료)",
            "isMatchedApplication()",
        ),
        # 파트너 탭 (필수)
        (
            "파트너 탭",
            "new",
            "신규견적",
            "아직 내 견적 미제출 콜",
            "탭=new AND 해당 기사 견적 없음",
            "견적 제출 또는 마감",
            "제휴기사",
            "applications; driver_quotes",
            "partner tab; application_id",
            "파트너",
            "제출견적",
            "PARTNER_DASHBOARD_TABS",
        ),
        (
            "파트너 탭",
            "quoted",
            "제출견적",
            "내가 견적 제출한 콜(미매칭)",
            "탭=quoted AND driver_quotes 존재 AND final_selected 아님",
            "매칭 성공 또는 마감",
            "제휴기사",
            "driver_quotes",
            "status; application_id",
            "파트너",
            "매칭성공",
            "",
        ),
        (
            "파트너 탭",
            "matched",
            "매칭성공",
            "내 견적이 최종 선택됨",
            "final_selected_quote_id = 내 견적 AND source=member",
            "운행 완료(서브필터)",
            "고객·시스템",
            "applications",
            "final_selected_quote_id; final_selected_quote_source",
            "파트너",
            "(종료)",
            "",
        ),
        # 스폰서 (필수)
        (
            "스폰서 탭",
            "review",
            "지원검토",
            "가승인·승인 대기 지원건",
            "sponsor_preapprovals.status=preapproved OR application.sponsor_support_status=preapproved",
            "지원확정(approved) 또는 거절",
            "스폰서",
            "sponsor_preapprovals; applications",
            "status; estimated_support_amount",
            "스폰서",
            "지원확정",
            "SPONSOR_MAIN_TABS review",
        ),
        (
            "스폰서 탭",
            "confirmed",
            "지원확정",
            "승인된 지원·지급 관리",
            "status=approved",
            "payout_status=completed 또는 취소",
            "스폰서",
            "sponsor_preapprovals",
            "status; approved_support_amount; payout_status",
            "스폰서",
            "지급중",
            "",
        ),
        (
            "스폰서 지급",
            "processing",
            "지급중",
            "승인 후 지급 진행",
            "payout_status=processing (승인 시 기본값)",
            "지급완료 처리",
            "스폰서",
            "sponsor_preapprovals",
            "payout_status",
            "스폰서",
            "지급완료",
            "sql/sponsor_preapprovals_payout.sql",
        ),
        (
            "스폰서 지급",
            "completed",
            "지급완료",
            "지급 완료 처리됨",
            "payout_status=completed",
            "(종료)",
            "스폰서",
            "sponsor_preapprovals",
            "payout_status",
            "스폰서",
            "",
            "",
        ),
        # 지원 UI 단계
        (
            "지원 표시단계",
            "planned",
            "지원검토",
            "예상·예정 지원금 표시 단계",
            "sponsor_support_status ≠ approved AND confirmed_total 없음",
            "sponsor_support_status=approved 또는 confirmed_total>0",
            "스폰서·시스템",
            "applications; driver_quotes",
            "sponsor_support_status; support_breakdown",
            "클라·파트너·어드민·스폰서",
            "지원확정",
            "buildQuoteSupportDisplayModel.resolveStage",
        ),
        (
            "지원 표시단계",
            "confirmed",
            "지원확정",
            "확정 지원·할인 적용가 표시",
            "approved 또는 confirmed_total_support>0",
            "(해당 신청 종료)",
            "스폰서",
            "driver_quotes; sponsor_preapprovals",
            "confirmed_*; approved_support_amount",
            "전체",
            "",
            "",
        ),
        # quote_status 상세
        (
            "견적 라이프사이클",
            "collecting",
            "견적요청중",
            "견적 수집 중",
            "quote_status=collecting",
            "마감 조건 충족",
            "시스템",
            "applications",
            "quote_status; quote_deadline_at",
            "어드민·클라이언트",
            "closed_by_*",
            "lib/quote-auction.ts",
        ),
        (
            "견적 라이프사이클",
            "auto_selected",
            "자동선정",
            "시스템 자동 선정(매칭 전)",
            "마감 후 auto_selected_quote_id 설정",
            "고객 final_selected 또는 수동",
            "시스템",
            "applications",
            "auto_selected_quote_id; quote_status",
            "어드민",
            "final_selected",
            "",
        ),
        (
            "견적 라이프사이클",
            "final_selected",
            "매칭완료(시스템값)",
            "최종 견적 확정",
            "고객 매칭 또는 자동확정",
            "contract_pending / completed",
            "고객·시스템",
            "applications",
            "final_selected_quote_id; quote_status",
            "전체",
            "completed",
            "",
        ),
        (
            "스폰서 집계",
            "preapproved",
            "지원검토(집계)",
            "가승인만 존재",
            "sponsor_support_status=preapproved",
            "approved 또는 rejected",
            "시스템",
            "applications",
            "sponsor_support_status",
            "어드민·스폰서",
            "approved|mixed",
            "",
        ),
        (
            "스폰서 집계",
            "approved",
            "지원확정(집계)",
            "승인된 지원 존재",
            "sponsor_support_status=approved",
            "rejected/mixed 전환",
            "스폰서",
            "applications",
            "sponsor_support_status",
            "전체",
            "",
            "admin-sponsor-confirmed",
        ),
        (
            "sponsor_preapprovals",
            "preapproved",
            "가승인",
            "자동/수동 매칭 직후",
            "INSERT preapproval status=preapproved",
            "approved/rejected/cancelled",
            "시스템·스폰서",
            "sponsor_preapprovals",
            "status",
            "스폰서",
            "approved",
            "",
        ),
        (
            "sponsor_preapprovals",
            "approved",
            "승인",
            "스폰서 지원 확정",
            "스폰서 지원확정 액션",
            "payout 완료 또는 취소",
            "스폰서",
            "sponsor_preapprovals",
            "status; approved_support_amount",
            "스폰서",
            "processing",
            "",
        ),
        (
            "sponsor_preapprovals",
            "rejected",
            "거절",
            "지원 거절",
            "스폰서 거절",
            "(종료)",
            "스폰서",
            "sponsor_preapprovals",
            "status",
            "스폰서",
            "",
            "",
        ),
        (
            "매칭 서브",
            "in_progress",
            "진행중",
            "매칭 후 출발 전",
            "matched 탭 AND 출발일시 > now",
            "출발일시 경과",
            "시스템",
            "applications",
            "departure_date; departure_time",
            "클라·파트너",
            "진행완료",
            "matchedRunStatus()",
        ),
        (
            "매칭 서브",
            "completed",
            "진행완료",
            "출발일 경과",
            "matched AND 출발일시 ≤ now",
            "(종료)",
            "시스템",
            "applications",
            "departure_date",
            "클라·파트너",
            "",
            "",
        ),
    ]
    cols = [
        "상태구분",
        "시스템 상태값",
        "한글 표시명",
        "의미",
        "진입 조건",
        "종료 조건",
        "변경 주체",
        "관련 테이블",
        "관련 필드",
        "표시 대시보드",
        "다음 상태",
        "주의사항",
    ]
    return pd.DataFrame(rows, columns=cols)


def build_calculation_rules() -> pd.DataFrame:
    rows = [
        (
            "totalPlannedSupport",
            "예상 지원금",
            "resolvePlannedTotalSupport(quote, ctx)",
            "planned_total_support; preapproved_support_amount; estimated_support_amount; sponsor 규칙 인원",
            "planned_total_support; support_breakdown.planned_total_support; totalPlannedSupport",
            "sponsor_quote_enabled=true AND 일반견적가 존재",
            "인원25, 인당2만, 상한80만",
            "500000",
            "resolvePlannedTotalSupport",
            "lib/support-calculation.ts",
            "전체 지원금 블록",
            "approved를 planned fallback으로 쓰지 않음(명시 소스 없을 때만)",
        ),
        (
            "totalConfirmedSupport",
            "확정 지원금",
            "resolveConfirmedTotalSupport(quote, options)",
            "confirmed_total_support; approved_support_amount; support_breakdown.totalConfirmedSupport; sponsor approved 합",
            "confirmed_total_support; totalConfirmedSupport",
            "스폰서 status=approved 또는 isConfirmed",
            "확정총 500000",
            "500000",
            "resolveConfirmedTotalSupport",
            "lib/support-calculation.ts",
            "지원확정 단계",
            "",
        ),
        (
            "customerPlannedSupport",
            "고객 예상 지원금",
            "resolvePlannedCustomerSupport(quote)",
            "planned_customer_support; customer_support_amount; client_reward_amount",
            "planned_customer_support; customerPlannedSupport",
            "지원 견적 제출 후",
            "고객예정 300000",
            "300000",
            "resolvePlannedCustomerSupport",
            "lib/support-calculation.ts",
            "지원검토 display_rows",
            "min(customer, total, normal_price)",
        ),
        (
            "customerConfirmedSupport",
            "고객 확정 지원금",
            "calculateSupportDistribution (client_priority) OR ratio OR derive",
            "planned_customer; confirmed_total; normal_price; final_discount_price; extension",
            "confirmed_customer_support; customerConfirmedSupport",
            "지원확정",
            "normal600000 discount300000 ext40000",
            "260000",
            "deriveCustomerConfirmedSupport; calculateSupportDistribution",
            "lib/support-calculation.ts",
            "지원확정",
            "normal - final_discount - extension (역산)",
        ),
        (
            "partnerPlannedSupport",
            "기사 예상 지원금",
            "calculatePlannedDriverSupport(total, customer, extension)",
            "planned_driver_support; driver_support_amount; total; customer; extension",
            "planned_driver_support; partnerPlannedSupport",
            "총 예정 존재",
            "total500000 customer300000 ext40000",
            "160000",
            "calculatePlannedDriverSupport",
            "lib/support-calculation.ts",
            "파트너 견적 폼",
            "total - customer - extension",
        ),
        (
            "partnerConfirmedSupport",
            "기사 확정 지원금",
            "confirmed_total - customer_confirmed - extension (또는 breakdown 저장값)",
            "confirmed_driver_support; final_driver_support_amount",
            "confirmed_driver_support; partnerConfirmedSupport",
            "지원확정",
            "total500000 customer260000 ext40000",
            "200000",
            "computeConfirmedFromPlanned; buildQuoteSupportDisplayModel",
            "lib/quote-support-display-model.ts",
            "지원확정",
            "연장 차감 후 기사 몫",
        ),
        (
            "plannedExtensionSupport",
            "예상 연장 지원금",
            "extensionPlannedFromPartnerSupport(partnerPlanned, extension_round)",
            "planned_driver_support; applications.extension_round",
            "planned_extension_support; support_breakdown.planned_extension_support",
            "extension_round > 0",
            "기사예정200000 round1",
            "40000",
            "extensionPlannedFromPartnerSupport",
            "lib/support-calculation.ts",
            "지원검토",
            "1회차 20% 2회차+ 40%",
        ),
        (
            "confirmedExtensionSupport",
            "확정 연장 지원금",
            "calculateExtension(partnerBeforeExt, extension_round) OR calculateExtensionSupport(partnerConfirmed)",
            "extension_support_amount; confirmed_extension_support; extension_round",
            "extension_support_amount; extensionSupport",
            "지원확정 AND extension_round>0",
            "기사확정200000",
            "40000",
            "calculateExtensionSupport; calculateExtension",
            "lib/support-calculation.ts; lib/quote-support-snapshot.ts",
            "지원확정",
            "DB 코멘트: partnerConfirmed×20%",
        ),
        (
            "supportDiscountPlannedPrice",
            "지원금 할인 예상가",
            "calculatePlannedDiscountPrice(normal, customerPlanned, extensionPlanned)",
            "normal_price; planned_customer_support; planned_extension_support",
            "planned_discount_price; member_price; supportDiscountPlannedPrice",
            "지원검토",
            "normal600000 customer300000 ext40000",
            "260000",
            "calculatePlannedDiscountPrice",
            "lib/support-calculation.ts",
            "지원검토·할인견적 선택",
            "max(normal - customer - extension, 0)",
        ),
        (
            "supportDiscountAppliedPrice",
            "지원금 할인 적용가",
            "max(normal - customerConfirmed - extensionConfirmed, 0)",
            "normal_price; confirmed_customer; confirmed_extension; final_member_price",
            "confirmed_discount_price; final_discount_price; supportDiscountAppliedPrice",
            "지원확정",
            "normal600000 customer260000 ext40000",
            "300000",
            "calculateSupportDiscountPrice; computeConfirmedFromPlanned",
            "lib/support-calculation.ts",
            "지원확정·매칭",
            "",
        ),
        (
            "selectedPrice",
            "선택 견적",
            "resolveEffectiveSelectedPriceType + 금액",
            "selected_price_type; selected_price; normal; planned/applied discount",
            "selected_price; applications.selected_price",
            "매칭 확정",
            "type=normal price600000",
            "600000",
            "resolveEffectiveSelectedPriceType; resolveSelectedPriceDisplay",
            "lib/selected-price-display.ts",
            "매칭 패널",
            "type 우선, 금액 추론 보조",
        ),
        (
            "finalPaymentPrice",
            "최종 결제가격",
            "선택 견적가 = 매칭 금액(동일)",
            "selected_price; selected_price_type",
            "selected_price",
            "매칭완료",
            "할인적용가300000",
            "300000",
            "buildQuoteSupportDisplayModel.selected_price",
            "lib/quote-support-display-model.ts; ClientMatchedPricePanel",
            "클라·파트너 매칭",
            "LABEL.finalPaymentPrice",
        ),
        (
            "settlementClientPriority",
            "고객 지원금 우선보장",
            "customerConfirmed = min(customerPlanned, totalConfirmed); partner = total - customer",
            "support_settlement_type=client_priority; planned_customer; confirmed_total",
            "confirmed_customer_support; confirmed_driver_support",
            "확정 시 settlement=client_priority",
            "planned고객30만 확정총50만",
            "고객30만 기사20만",
            "calculateSupportDistribution",
            "lib/support-calculation.ts",
            "파트너 정산모드",
            "고객 예정 우선, 잔액 기사",
        ),
        (
            "settlementRatio",
            "비율정산",
            "customerConfirmed = round(customerPlanned * (confirmedTotal / plannedTotal))",
            "support_settlement_type=ratio; planned 비율",
            "confirmed_customer_support",
            "확정 시 settlement=ratio",
            "planned 30/50 확정총40만",
            "고객24만 기사16만",
            "calculateSupportDistribution",
            "lib/support-calculation.ts",
            "파트너 정산모드",
            "예정 비율 유지",
        ),
        (
            "ruleTotalPlanned",
            "예상 지원금(규칙산출)",
            "calculateTotalPlannedSupport(passengers, per_person, per_case, max)",
            "sponsor_rules; passenger_count",
            "estimated_support_amount; preapproval.estimated_support_amount",
            "스폰서 매칭·견적 전",
            "25명×2만",
            "500000",
            "calculateTotalPlannedSupport",
            "lib/support-calculation.ts",
            "스폰서 매칭",
            "일일예산·최대인원 cap",
        ),
    ]
    cols = [
        "계산항목",
        "한글명",
        "계산식",
        "입력 필드",
        "출력 필드",
        "적용 조건",
        "예시 입력값",
        "예시 결과값",
        "사용하는 함수",
        "사용하는 파일",
        "표시 위치",
        "주의사항",
    ]
    return pd.DataFrame(rows, columns=cols)


def build_dashboard_display_rules() -> pd.DataFrame:
    rows = [
        # 클라이언트
        (
            "클라이언트 대시보드",
            "견적요청중",
            "일반견적가",
            "일반견적가",
            "견적 카드 항상(제휴/일반기사)",
            "",
            "driver_quotes.price; guest_driver_quotes.price",
            "normal_price",
            "price -> target_normal_price",
            "",
            "ClientApplicationListItem; client-display",
            "",
        ),
        (
            "클라이언트 대시보드",
            "견적요청중",
            "지원금 할인 예상가",
            "지원금 할인 예상가",
            "sponsor_quote_enabled AND 지원검토(지원확정 아님)",
            "selected_price_type=normal ONLY on matched panel",
            "support_discount_planned_price; member_price; breakdown",
            "supportDiscountPlannedPrice",
            "buildQuoteSupportDisplayModel",
            "매칭완료(할인 선택)",
            "app/client/dashboard; SupportQuoteBreakdown",
            "지원검토 시 예정가",
        ),
        (
            "클라이언트 대시보드",
            "견적요청중",
            "지원금 할인 적용가",
            "지원금 할인 적용가",
            "sponsor_support_status=approved OR quote approved",
            "지원검토 단계",
            "support_discount_applied_price; final_member_price",
            "supportDiscountAppliedPrice",
            "buildQuoteSupportDisplayModel",
            "",
            "client-display.ts",
            "지원확정 시만",
        ),
        (
            "클라이언트 대시보드",
            "매칭완료",
            "선택 견적",
            "매칭견적가; 선택 견적",
            "final_selected_quote_id 존재",
            "",
            "selected_price_type; selected_price; selected_price_label",
            "selected_price",
            "selected_price_type -> selected_price",
            "매칭완료",
            "ClientMatchedPricePanel",
            "type=normal이면 일반가",
        ),
        (
            "클라이언트 대시보드",
            "매칭완료",
            "지원금 블록",
            "(지원 단계 행 전체)",
            "selected_price_type ≠ normal (할인견적 매칭)",
            "selected_price_type=normal (일반견적 매칭)",
            "display_rows; support_stage",
            "buildQuoteSupportDisplayModel",
            "show_normal_price=false when normal selected",
            "",
            "ClientMatchedPricePanel; quote-support-display-model",
            "일반견적 선택 시 지원금 항목 숨김",
        ),
        (
            "클라이언트 대시보드",
            "매칭완료",
            "최종 결제가격",
            "최종 결제가격",
            "매칭완료 탭",
            "",
            "selected_price",
            "selected_price",
            "",
            "",
            "ClientMatchedPricePanel",
            "",
        ),
        # 파트너
        (
            "파트너 대시보드",
            "신규견적/제출견적",
            "예상 지원금",
            "총 예상 지원금",
            "sponsor_quote_enabled",
            "일반견적만 제출",
            "planned_total_support; breakdown",
            "totalPlannedSupport",
            "preapproved -> planned",
            "견적제출",
            "PartnerCallExpandPanel; PartnerSupportSummary",
            "",
        ),
        (
            "파트너 대시보드",
            "매칭성공",
            "선택 견적",
            "선택 견적",
            "final_selected = 내 견적",
            "",
            "selected_price; selected_price_type",
            "admin/partner resolution",
            "selected_price_type 우선",
            "",
            "PartnerMatchedPricePanel",
            "",
        ),
        (
            "파트너 대시보드",
            "견적 폼",
            "정산모드",
            "고객 지원금 우선보장 / 비율정산",
            "지원금 견적 제출 시",
            "",
            "support_settlement_type",
            "SETTLEMENT_OPTIONS",
            "",
            "견적제출·수정",
            "partner/dashboard; SETTLEMENT_OPTIONS",
            "",
        ),
        (
            "파트너 대시보드",
            "콜 카드",
            "연장 회차",
            "연장 회차",
            "extension_round > 0",
            "",
            "applications.extension_round",
            "extension_count",
            "",
            "",
            "PartnerCallCard",
            "",
        ),
        # 스폰서
        (
            "스폰서 대시보드",
            "지원검토",
            "예상 지원금",
            "예상 지원금",
            "tab=review",
            "tab=confirmed",
            "estimated_support_amount",
            "planned_total",
            "preapproval",
            "지원확정",
            "SponsorCallCard",
            "",
        ),
        (
            "스폰서 대시보드",
            "지원확정",
            "확정 지원금",
            "확정 지원금",
            "status=approved",
            "",
            "approved_support_amount",
            "confirmed_total",
            "",
            "지원금입력",
            "SponsorCallCard",
            "매칭 후 readonly",
        ),
        (
            "스폰서 대시보드",
            "지원확정",
            "지급상태",
            "지급중/지급완료",
            "approved",
            "",
            "payout_status",
            "",
            "",
            "지급완료 처리",
            "SponsorCallCard",
            "",
        ),
        # 어드민
        (
            "어드민 대시보드",
            "신청 상세",
            "제휴기사 견적 카드",
            "일반견적가·지원금 행",
            "member quote 존재",
            "",
            "support_breakdown; planned_*; confirmed_*",
            "buildQuoteSupportDisplayModel",
            "application sponsor approved",
            "",
            "ApplicationDetailMatchedPanel; admin-member-quote-support-display",
            "selected_price_type 우선",
        ),
        (
            "어드민 대시보드",
            "신청 상세",
            "선택 견적",
            "선택 견적",
            "final_selected 존재",
            "",
            "selected_price_type; selected_price",
            "resolveAdminSelectedQuoteDisplay",
            "support_breakdown",
            "",
            "lib/admin-selected-quote-price.ts",
            "할인가 오표시 방지",
        ),
        (
            "어드민 대시보드",
            "신청 상세",
            "지원 단계",
            "지원검토/지원확정",
            "sponsor_confirmed",
            "",
            "sponsor_support_status; support_breakdown.isConfirmed",
            "support_stage",
            "admin-sponsor-confirmed",
            "",
            "admin-member-quote-support-display",
            "",
        ),
        # 공통 선택견적 규칙
        (
            "공통",
            "매칭/상세",
            "지원금 할인 예상가",
            "지원금 할인 예정가",
            "support_stage=지원검토 AND 할인견적 선택",
            "일반견적 선택 OR 지원확정(적용가로 대체)",
            "planned_discount_price; support_stage",
            "final_discount when confirmed",
            "selected_price_type=support_planned",
            "",
            "lib/quote-support-display-model.ts",
            "지원검토=예정가 라벨",
        ),
        (
            "공통",
            "매칭/상세",
            "지원금 할인 적용가",
            "지원금 할인 적용가",
            "support_stage=지원확정 AND 할인견적 선택",
            "일반견적 선택",
            "final_discount_price; confirmed_discount_price",
            "supportDiscountAppliedPrice",
            "selected_price_type=support_confirmed",
            "",
            "lib/selected-price-display.ts",
            "지원확정=적용가 라벨",
        ),
        (
            "공통",
            "매칭/상세",
            "일반견적가 단독",
            "일반견적가",
            "selected_price_type=normal OR effectiveType=normal",
            "할인견적 매칭",
            "price; selected_price",
            "normal_price",
            "",
            "",
            "전 대시보드",
            "지원금 관련 행 숨김(show_normal_price 등)",
        ),
    ]
    cols = [
        "대시보드",
        "화면/탭",
        "표시 항목",
        "표시 라벨",
        "표시 조건",
        "숨김 조건",
        "사용하는 필드",
        "사용하는 계산값",
        "fallback",
        "버튼/액션",
        "관련 컴포넌트",
        "주의사항",
    ]
    return pd.DataFrame(rows, columns=cols)


def main() -> int:
    outputs: list[tuple[Path, str, pd.DataFrame, list[str]]] = []

    df_field = build_field_map()
    verify_dataframe_korean(df_field, ["한글명", "역할정의"])
    outputs.append((FIELD_MAP_XLSX, "field_map", df_field, ["한글명", "필드명", "테이블명"]))

    df_state = build_state_definition()
    verify_dataframe_korean(df_state, ["한글 표시명", "의미"])
    outputs.append((STATE_XLSX, "state_definition", df_state, ["한글 표시명", "상태구분", "시스템 상태값"]))

    df_calc = build_calculation_rules()
    verify_dataframe_korean(df_calc, ["한글명", "계산식"])
    outputs.append((CALC_XLSX, "calculation_rules", df_calc, ["한글명", "계산항목", "계산식"]))

    df_display = build_dashboard_display_rules()
    verify_dataframe_korean(df_display, ["표시 라벨", "표시 항목"])
    outputs.append((DISPLAY_XLSX, "dashboard_display_rules", df_display, ["대시보드", "표시 라벨", "화면/탭"]))

    print("Generated files:")
    for path, sheet, df, preview_cols in outputs:
        saved = save_workbook(path, sheet, df)
        print(f"  {path}")
        print_verify(path, sheet, saved, preview_cols)

    return 0


if __name__ == "__main__":
    sys.exit(main())
